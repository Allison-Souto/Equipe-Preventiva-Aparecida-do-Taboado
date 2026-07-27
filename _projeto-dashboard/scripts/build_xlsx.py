import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

FONT = 'Arial'
HEADER_FILL = PatternFill('solid', fgColor='1F2937')
HEADER_FONT = Font(name=FONT, size=10, bold=True, color='FFFFFF')
TITLE_FONT = Font(name=FONT, size=14, bold=True, color='1F2937')
SUB_FONT = Font(name=FONT, size=9, italic=True, color='6B7280')
BASE_FONT = Font(name=FONT, size=10)
BOLD_FONT = Font(name=FONT, size=10, bold=True)
THIN = Side(style='thin', color='D1D5DB')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GREEN = Font(name=FONT, size=10, color='0E7C3A')
RED = Font(name=FONT, size=10, color='B3261E')
CUR_FMT = 'R$ #,##0.00'
PCT_FMT = '0.0%'


def safe_text(v):
    """Prevent spreadsheet apps from misreading text starting with =,+,-,@ as a formula."""
    if isinstance(v, str) and v[:1] in ('=', '+', '-', '@'):
        return '​' + v
    return v

def style_header(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col+ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 28

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

d = json.load(open('/sessions/modest-vigilant-hypatia/mnt/outputs/parsed.json', encoding='utf-8'))
data = json.load(open('/sessions/modest-vigilant-hypatia/mnt/outputs/dashboard_data.json', encoding='utf-8'))

wb = Workbook()

# ============ SHEET 1: Resumo (Totais por medicao) ============
ws = wb.active
ws.title = 'Resumo'
ws['A1'] = 'Contrato 131/2024 — Equipe Preventiva (Solo Construções e Incorporações)'
ws['A1'].font = TITLE_FONT
ws['A2'] = 'Totais por medição — valores em R$. Colunas de variação calculadas por fórmula.'
ws['A2'].font = SUB_FONT
headers = ['Medição', 'Mês', 'Data do boletim', 'Período', 'Mão de obra', 'Materiais', 'Custo operacional', 'Total', 'Variação vs mês anterior', 'Variação %']
row0 = 4
for j, h in enumerate(headers, start=1):
    ws.cell(row=row0, column=j, value=h)
style_header(ws, row0, len(headers))

totais = data['totais']
first_data_row = row0 + 1
for i, t in enumerate(totais):
    r = first_data_row + i
    ws.cell(row=r, column=1, value=t['medicao'])
    ws.cell(row=r, column=2, value=t['mes_label'])
    ws.cell(row=r, column=3, value=t['data'])
    ws.cell(row=r, column=4, value=t['periodo'])
    ws.cell(row=r, column=5, value=t['mao_de_obra'])
    ws.cell(row=r, column=6, value=t['materiais'])
    ws.cell(row=r, column=7, value=t['custo_operacional'])
    ws.cell(row=r, column=8, value=f'=E{r}+F{r}+G{r}')
    if i == 0:
        ws.cell(row=r, column=9, value=None)
        ws.cell(row=r, column=10, value=None)
    else:
        ws.cell(row=r, column=9, value=f'=H{r}-H{r-1}')
        ws.cell(row=r, column=10, value=f'=IF(H{r-1}=0,"",(H{r}-H{r-1})/H{r-1})')
    for col in (5, 6, 7, 8, 9):
        ws.cell(row=r, column=col).number_format = CUR_FMT
    ws.cell(row=r, column=10).number_format = PCT_FMT
    for c in range(1, 11):
        ws.cell(row=r, column=c).font = BASE_FONT
        ws.cell(row=r, column=c).border = BORDER

last_row = first_data_row + len(totais) - 1
tot_r = last_row + 2
ws.cell(row=tot_r, column=4, value='TOTAL DO CONTRATO').font = BOLD_FONT
for col, letter in [(5,'E'),(6,'F'),(7,'G'),(8,'H')]:
    ws.cell(row=tot_r, column=col, value=f'=SUM({letter}{first_data_row}:{letter}{last_row})')
    ws.cell(row=tot_r, column=col).font = BOLD_FONT
    ws.cell(row=tot_r, column=col).number_format = CUR_FMT

ws.conditional_formatting.add(f'I{first_data_row}:I{last_row}',
    CellIsRule(operator='greaterThan', formula=['0'], font=GREEN))
ws.conditional_formatting.add(f'I{first_data_row}:I{last_row}',
    CellIsRule(operator='lessThan', formula=['0'], font=RED))

autosize(ws, [10, 8, 14, 22, 14, 14, 16, 14, 18, 12])
ws.freeze_panes = f'A{first_data_row}'

# ============ SHEET 2: Mao de Obra (evolucao por profissional) ============
ws2 = wb.create_sheet('Mão de Obra')
ws2['A1'] = 'Mão de obra — carga horária e valores por profissional, por medição'
ws2['A1'].font = TITLE_FONT
ws2['A2'] = 'Delta = variação em relação à medição anterior (valor pré-calculado a partir dos boletins). Vazio = sem valor na medição.'
ws2['A2'].font = SUB_FONT

meses = data['meses']
row0 = 4
ws2.cell(row=row0, column=1, value='Profissional')
ws2.cell(row=row0-1, column=1, value='')
col = 2
month_cols = {}
for m in meses:
    ws2.cell(row=row0, column=col, value=m + ' — Horas')
    ws2.cell(row=row0, column=col+1, value=m + ' — Valor (R$)')
    ws2.cell(row=row0, column=col+2, value=m + ' — Δ Valor')
    month_cols[m] = col
    col += 3
ncols = col - 1
style_header(ws2, row0, ncols)

mo_deltas_json = data['mao_de_obra_deltas']
mo_valor = data['mao_de_obra']['valor']
mo_qtd = data['mao_de_obra']['qtd']
r = row0 + 1
for role in mo_valor.keys():
    ws2.cell(row=r, column=1, value=safe_text(role.title()))
    ws2.cell(row=r, column=1).font = BOLD_FONT
    for idx, m in enumerate(meses):
        c = month_cols[m]
        qtd = mo_qtd[role][idx]
        val = mo_valor[role][idx]
        ws2.cell(row=r, column=c, value=qtd if qtd is not None else None)
        ws2.cell(row=r, column=c+1, value=val if val is not None else None)
        d_entry = mo_deltas_json.get(role, [None]*len(meses))[idx]
        dv = d_entry.get('delta_valor') if d_entry else None
        ws2.cell(row=r, column=c+2, value=dv)
        ws2.cell(row=r, column=c+1).number_format = CUR_FMT
        ws2.cell(row=r, column=c+2).number_format = CUR_FMT
        for cc in (c, c+1, c+2):
            ws2.cell(row=r, column=cc).font = BASE_FONT
            ws2.cell(row=r, column=cc).border = BORDER
    ws2.cell(row=r, column=1).border = BORDER
    r += 1

last_r2 = r - 1
for m in meses:
    c = month_cols[m]
    delta_letter = get_column_letter(c+2)
    ws2.conditional_formatting.add(f'{delta_letter}{row0+1}:{delta_letter}{last_r2}',
        CellIsRule(operator='greaterThan', formula=['0'], font=GREEN))
    ws2.conditional_formatting.add(f'{delta_letter}{row0+1}:{delta_letter}{last_r2}',
        CellIsRule(operator='lessThan', formula=['0'], font=RED))

widths2 = [30] + [12,14,12]*len(meses)
autosize(ws2, widths2)
ws2.freeze_panes = ws2.cell(row=row0+1, column=2).coordinate

# ============ SHEET 3: Custo Operacional (mesma estrutura) ============
ws3 = wb.create_sheet('Custo Operacional')
ws3['A1'] = 'Custo operacional — quantidade e valores por item, por medição'
ws3['A1'].font = TITLE_FONT
ws3['A2'] = 'Delta = variação em relação à medição anterior (valor pré-calculado a partir dos boletins).'
ws3['A2'].font = SUB_FONT
row0 = 4
ws3.cell(row=row0, column=1, value='Item')
col = 2
month_cols3 = {}
for m in meses:
    ws3.cell(row=row0, column=col, value=m + ' — Qtd')
    ws3.cell(row=row0, column=col+1, value=m + ' — Valor (R$)')
    ws3.cell(row=row0, column=col+2, value=m + ' — Δ Valor')
    month_cols3[m] = col
    col += 3
ncols3 = col - 1
style_header(ws3, row0, ncols3)

co_deltas_json = data['custo_operacional_deltas']
co_valor = data['custo_operacional']['valor']
co_qtd = data['custo_operacional']['qtd']
r = row0 + 1
for item in co_valor.keys():
    ws3.cell(row=r, column=1, value=safe_text(item.title()))
    ws3.cell(row=r, column=1).font = BOLD_FONT
    for idx, m in enumerate(meses):
        c = month_cols3[m]
        qtd = co_qtd[item][idx]
        val = co_valor[item][idx]
        ws3.cell(row=r, column=c, value=qtd if qtd is not None else None)
        ws3.cell(row=r, column=c+1, value=val if val is not None else None)
        d_entry = co_deltas_json.get(item, [None]*len(meses))[idx]
        dv = d_entry.get('delta_valor') if d_entry else None
        ws3.cell(row=r, column=c+2, value=dv)
        ws3.cell(row=r, column=c+1).number_format = CUR_FMT
        ws3.cell(row=r, column=c+2).number_format = CUR_FMT
        for cc in (c, c+1, c+2):
            ws3.cell(row=r, column=cc).font = BASE_FONT
            ws3.cell(row=r, column=cc).border = BORDER
    ws3.cell(row=r, column=1).border = BORDER
    r += 1
last_r3 = r - 1
for m in meses:
    c = month_cols3[m]
    delta_letter = get_column_letter(c+2)
    ws3.conditional_formatting.add(f'{delta_letter}{row0+1}:{delta_letter}{last_r3}',
        CellIsRule(operator='greaterThan', formula=['0'], font=GREEN))
    ws3.conditional_formatting.add(f'{delta_letter}{row0+1}:{delta_letter}{last_r3}',
        CellIsRule(operator='lessThan', formula=['0'], font=RED))
widths3 = [45] + [10,14,12]*len(meses)
autosize(ws3, widths3)
ws3.freeze_panes = ws3.cell(row=row0+1, column=2).coordinate

# ============ SHEET 3b: Materiais - Recorrentes (mesma estrutura de Mao de Obra / Custo Operacional) ============
ws3b = wb.create_sheet('Materiais - Recorrentes')
ws3b['A1'] = 'Materiais e insumos recorrentes — quantidade e valores por item, por medição'
ws3b['A1'].font = TITLE_FONT
ws3b['A2'] = 'Apenas insumos comprados em 2 ou mais medições (evolução mensal com sentido). Itens de compra única estão na aba "Curva ABC Acumulada". Delta = valor pré-calculado a partir dos boletins.'
ws3b['A2'].font = SUB_FONT
row0 = 4
ws3b.cell(row=row0, column=1, value='Insumo')
col = 2
month_cols3b = {}
for m in meses:
    ws3b.cell(row=row0, column=col, value=m + ' — Qtd')
    ws3b.cell(row=row0, column=col+1, value=m + ' — Valor (R$)')
    ws3b.cell(row=row0, column=col+2, value=m + ' — Δ Valor')
    month_cols3b[m] = col
    col += 3
ncols3b = col - 1
style_header(ws3b, row0, ncols3b)

mat_deltas_json = data['materiais_recorrentes_deltas']
mat_valor = data['materiais_recorrentes']['valor']
mat_qtd = data['materiais_recorrentes']['qtd']
r = row0 + 1
for item in mat_valor.keys():
    ws3b.cell(row=r, column=1, value=safe_text(item))
    ws3b.cell(row=r, column=1).font = BOLD_FONT
    for idx, m in enumerate(meses):
        c = month_cols3b[m]
        qtd = mat_qtd[item][idx]
        val = mat_valor[item][idx]
        ws3b.cell(row=r, column=c, value=qtd if qtd is not None else None)
        ws3b.cell(row=r, column=c+1, value=val if val is not None else None)
        d_entry = mat_deltas_json.get(item, [None]*len(meses))[idx]
        dv = d_entry.get('delta_valor') if d_entry else None
        ws3b.cell(row=r, column=c+2, value=dv)
        ws3b.cell(row=r, column=c+1).number_format = CUR_FMT
        ws3b.cell(row=r, column=c+2).number_format = CUR_FMT
        for cc in (c, c+1, c+2):
            ws3b.cell(row=r, column=cc).font = BASE_FONT
            ws3b.cell(row=r, column=cc).border = BORDER
    ws3b.cell(row=r, column=1).border = BORDER
    r += 1
last_r3b = r - 1
for m in meses:
    c = month_cols3b[m]
    delta_letter = get_column_letter(c+2)
    ws3b.conditional_formatting.add(f'{delta_letter}{row0+1}:{delta_letter}{last_r3b}',
        CellIsRule(operator='greaterThan', formula=['0'], font=GREEN))
    ws3b.conditional_formatting.add(f'{delta_letter}{row0+1}:{delta_letter}{last_r3b}',
        CellIsRule(operator='lessThan', formula=['0'], font=RED))
widths3b = [55] + [10,14,12]*len(meses)
autosize(ws3b, widths3b)
ws3b.freeze_panes = ws3b.cell(row=row0+1, column=2).coordinate

# ============ SHEET 4: Curva ABC Acumulada ============
ws4 = wb.create_sheet('Curva ABC Acumulada')
ws4['A1'] = 'Curva ABC de insumos — acumulado de todo o contrato (13 medições)'
ws4['A1'].font = TITLE_FONT
ws4['A2'] = 'Classe A = itens que somam até 80% do valor total; B = até 95%; C = os demais. % e classe pré-calculados a partir dos boletins.'
ws4['A2'].font = SUB_FONT
headers4 = ['#', 'Insumo', 'Valor total (R$)', 'Quantidade total', 'Unidade', 'Nº de medições em que aparece', '% do total', '% acumulado', 'Classe']
row0 = 4
for j, h in enumerate(headers4, start=1):
    ws4.cell(row=row0, column=j, value=h)
style_header(ws4, row0, len(headers4))

acc = data['abc_acumulado']
first_r4 = row0 + 1
last_r4 = first_r4 + len(acc) - 1
for i, item in enumerate(acc):
    r = first_r4 + i
    ws4.cell(row=r, column=1, value=i+1)
    ws4.cell(row=r, column=2, value=safe_text(item['desc_norm']))
    ws4.cell(row=r, column=3, value=item['valor_total'])
    ws4.cell(row=r, column=4, value=item['qtd_total'])
    ws4.cell(row=r, column=5, value=item['unidade'])
    ws4.cell(row=r, column=6, value=item['n_medicoes'])
    ws4.cell(row=r, column=7, value=item['pct'])
    ws4.cell(row=r, column=8, value=item['cum_pct'])
    ws4.cell(row=r, column=9, value=item['classe'])
    ws4.cell(row=r, column=3).number_format = CUR_FMT
    ws4.cell(row=r, column=7).number_format = PCT_FMT
    ws4.cell(row=r, column=8).number_format = PCT_FMT
    for c in range(1, 10):
        ws4.cell(row=r, column=c).font = BASE_FONT
        ws4.cell(row=r, column=c).border = BORDER

ws4.conditional_formatting.add(f'I{first_r4}:I{last_r4}',
    CellIsRule(operator='equal', formula=['"A"'], fill=PatternFill('solid', fgColor='D1F5DD'), font=GREEN))
ws4.conditional_formatting.add(f'I{first_r4}:I{last_r4}',
    CellIsRule(operator='equal', formula=['"B"'], fill=PatternFill('solid', fgColor='FDF0D0')))
autosize(ws4, [6, 60, 16, 14, 10, 14, 12, 12, 9])
ws4.freeze_panes = f'C{first_r4}'

# ============ SHEET 5-7: Raw detail (Materiais, Mao de Obra, Custo Operacional) ============
def write_detail_sheet(name, csv_path, col_widths):
    df = pd.read_csv(csv_path)
    ws = wb.create_sheet(name)
    ws['A1'] = name + ' — dados brutos extraídos dos 13 boletins de medição'
    ws['A1'].font = TITLE_FONT
    row0 = 3
    for j, h in enumerate(df.columns, start=1):
        ws.cell(row=row0, column=j, value=h)
    style_header(ws, row0, len(df.columns))
    for i, row in enumerate(df.itertuples(index=False)):
        r = row0 + 1 + i
        for j, val in enumerate(row, start=1):
            v = None if pd.isna(val) else val
            v = safe_text(v) if isinstance(v, str) else v
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = BASE_FONT
            if df.columns[j-1] in ('valor_periodo', 'preco_unitario', 'valor_contrato', 'valor_saldo', 'valor', 'qtd_periodo', 'qtd_contrato'):
                if 'valor' in df.columns[j-1] or 'preco' in df.columns[j-1]:
                    cell.number_format = CUR_FMT
    autosize(ws, col_widths)
    ws.freeze_panes = f'A{row0+1}'
    return ws

write_detail_sheet('Materiais - Detalhe', '/sessions/modest-vigilant-hypatia/mnt/outputs/materiais_detalhe.csv',
    [10, 8, 10, 60, 10, 10, 12, 12, 12])
write_detail_sheet('Mão de Obra - Detalhe', '/sessions/modest-vigilant-hypatia/mnt/outputs/mao_de_obra_detalhe.csv',
    [10, 8, 30, 45, 8, 12, 12, 12])
write_detail_sheet('Custo Operacional - Detalhe', '/sessions/modest-vigilant-hypatia/mnt/outputs/custo_operacional_detalhe.csv',
    [10, 8, 40, 45, 8, 12, 12, 12])

# reorder sheets: Resumo, Mão de Obra, Custo Operacional, Curva ABC, then detail sheets
order = ['Resumo', 'Mão de Obra', 'Custo Operacional', 'Materiais - Recorrentes', 'Curva ABC Acumulada',
         'Materiais - Detalhe', 'Mão de Obra - Detalhe', 'Custo Operacional - Detalhe']
wb._sheets = [wb[name] for name in order]
for ws_ in wb._sheets:
    ws_.sheet_view.showGridLines = False

out_path = '/sessions/modest-vigilant-hypatia/mnt/outputs/Medicoes_Equipe_Preventiva.xlsx'
wb.save(out_path)
print('saved', out_path)
